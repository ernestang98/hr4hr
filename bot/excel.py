from openpyxl import load_workbook
from datetime import date

filename_boss = "excel-boss.xlsx"

filename_customer = "excel-customer.xlsx"

workbook_boss = load_workbook(filename=filename_boss)
workbook_customer = load_workbook(filename=filename_customer)

excel_writer_time_dict = {
    "0900": "2",
    "1000": "3",
    "1100": "4",
    "1200": "5",
    "1300": "6",
    "1400": "7",
    "1500": "8",
    "1600": "9",
    "1700": "10",
    "1800": "11",
    "1900": "12",
    "2000": "13",
    "2100": "14",
    "2200": "15",
    "2300": "16",
}

alphabet_dict = {
    "SHIT": "0",
    "A": "1",
    "B": "2",
    "C": "3",
    "D": "4",
    "E": "5",
    "F": "6",
    "G": "7",
    "H": "8",
    "I": "9",
    "J": "10",
    "K": "11",
    "L": "12",
    "M": "13",
    "N": "14",
    "O": "15",
    "P": "16",
    "Q": "17",
    "R": "18",
    "S": "19",
    "T": "20",
    "U": "21",
    "V": "22",
    "W": "23",
    "X": "24",
    "Y": "25",
    "Z": "26",
}

inverted_alphabet_dict = {v: k for (k, v) in alphabet_dict.items()}


def convertstringtodate(the_string):
    split = the_string.split("/")
    return date(int(split[2]), int(split[1]), int(split[0]))


def getcell(time_input, date_input):
    try:
        f_date = convertstringtodate("16/10/21")
        l_date = convertstringtodate(date_input)
        delta = l_date - f_date
        return str(getcoll(int(delta.days))) + str(excel_writer_time_dict[time_input])
    except:
        return "ERROR, out of possible range, if I am not mistaken it is 700 days from 16/10/21"

def getcellonlydate(date_input):
    try:
        f_date = convertstringtodate("16/10/21")
        l_date = convertstringtodate(date_input)
        delta = l_date - f_date
        return str(getcoll(int(delta.days)))
    except:
        return "ERROR, out of possible range, if I am not mistaken it is 700 days from 16/10/21"

def getcolmaxsevenhundred(number_of_days_infront):
    if number_of_days_infront <= 26:
        return inverted_alphabet_dict[str(number_of_days_infront)]
    else:
        if number_of_days_infront % 26 == 0:
            how_many_twenty_sixes = int(number_of_days_infront / 26) - 1
        else:
            how_many_twenty_sixes = number_of_days_infront // 26

        number_of_days_infront -= 26
        return inverted_alphabet_dict[str(how_many_twenty_sixes)] + getcol(number_of_days_infront)


def getcol(number_of_days_infront):
    temp = getcolmaxsevenhundred(number_of_days_infront)
    if len(temp) == 1:
        return temp
    else:
        return temp[0] + temp[len(temp)-1]


def getcoll(lol):
    return getcol(lol + 2)



def getsheet(val):
    if val == "Bowie Chan":
        return workbook_customer["Bowie Chan"], workbook_boss["Bowie Chan"]
    elif val == "John Boo":
        return workbook_customer["John Boo"], workbook_boss["Bowie Chan"]
    elif val == "Lawrence":
        return workbook_customer["Lawrence"], workbook_boss["Lawrence"]
    elif val == "Claudia Tong":
        return workbook_customer["Claudia Tong"], workbook_boss["Claudia Tong"]
    elif val == "Jess Zhang":
        return workbook_customer["Jess Zhang"], workbook_boss["Jess Zhang"]
    else:
        return None


def checktimeavailfordate(date, sheet):
    try:
        the_list = ["0900", "1000", "1100", "1200", "1300", "1400", "1500", "1600", "1700", "1800"]
        empty = []
        date = date.replace("2021", "21")
        for i in range(len(the_list)):
            if (sheet[str(str(getcellonlydate(date)) + str(i + 2))].value) == None:
                empty.append(the_list[i])

        return empty
    except:
        return "Error"


def makebooking(time, date, sheet1, sheet2, text):
    print(getcell(time, date.replace("2021", "21")))
    sheet1[str(getcell(time, date.replace("2021", "21")))] = "BOOKING"
    sheet2[str(getcell(time, date.replace("2021", "21")))] = text
    workbook_boss.save(filename=filename_boss)
    workbook_customer.save(filename=filename_customer)
        






# i think the max possible range is like ~700 days
# https://www.w3resource.com/python-exercises/python-basic-exercise-14.php
# https://stackoverflow.com/questions/23527887/getting-sheet-names-from-openpyxl
# https://realpython.com/openpyxl-excel-spreadsheets-python/





